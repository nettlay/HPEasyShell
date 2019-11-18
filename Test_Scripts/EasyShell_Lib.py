import platform
import shutil
import sys
import traceback
from Library.CommonLib import QAUtils, TxtUtils, getElementByType, Reg_Utils, YmlUtils
import os
import openpyxl


def setup():
    """
    Initial test environment before test
    include :
    copy this folder to c:\svc\hpeasyshell
    install services
    disable firewall
    disable UAC
    reboot
    """
    if sys.argv[0].upper() == 'RUN.EXE':
        self_path = os.getcwd()
    else:
        self_path = os.path.dirname(sys.argv[0])
    print(self_path)
    # ========================================
    # setup
    # ========================================
    try:
        if not os.path.exists(r'c:\svc'):
            os.mkdir(r'c:\svc')
        if os.path.exists(r'c:\svc\hpeasyshell'):
            if os.path.normcase(self_path) == os.path.normcase(r'c:\svc\hpeasyshell'):
                print(os.path.normcase(self_path), "====r'c:\svc\hpeasyshell'")
                if not os.path.exists(r'c:\svc\hpeasyshell\test_data\testset.xlsx'):
                    shutil.copy(r'c:\svc\hpeasyshell\test_data\testsetbak.xlsx',
                                r'c:\svc\hpeasyshell\test_data\testset.xlsx')
                return
            else:
                shutil.rmtree(r'c:\svc\hpeasyshell')
        try:
            shutil.copytree(self_path, r"C:\svc\hpeasyshell")
            shutil.copy(r'c:\svc\hpeasyshell\services\svcconfig.ini', r'c:\svc\svcconfig.ini')
            shutil.copy(r'c:\svc\hpeasyshell\services\runappasService.exe', r'c:\svc\runappasService.exe')
        except:
            print("May be some file exist or running, can not be replace, so skip this copy")
        os.system(r'c:\svc\runappasService.exe --startup auto install')
        # ----create test plan excel from scripts.yml
        if os.path.exists(r'c:\svc\hpeasyshell\test_data\script.yml'):
            yml = YmlUtils(r'c:\svc\hpeasyshell\test_data\script.yml')
            items = yml.get_item()
            wb = openpyxl.Workbook()
            sheet = wb.create_sheet(index=0, title='Test_Suite')
            sheet.cell(1, 1).value = "TestName"
            sheet.cell(1, 2).value = "TestResult"
            for index in range(len(items)):
                sheet.cell(index + 2, 1).value = list(items[index].keys())[0]
            wb.save(r'c:\svc\hpeasyshell\test_data\testset.xlsx')
        else:
            if not os.path.exists(r'c:\svc\hpeasyshell\test_data\testset.xlsx'):
                shutil.copy(r'c:\svc\hpeasyshell\test_data\testsetbak.xlsx', r'c:\svc\hpeasyshell\test_data\testset.xlsx')
        # ---UAC----
        reg = Reg_Utils()
        key = reg.open(r'SOFTWARE\Microsoft\Windows\CurrentVersion\Policies\System')
        reg.create_value(key=key, valueName='EnableLUA', regType=1, content=0)
        reg.close(key)
        # ---Firewall----
        os.system('netsh advfirewall set publicprofile state off')
        os.system('netsh advfirewall set privateprofile state off')
        os.system('shutdown -r -t 5')
    except:
        with open(r'c:\svc\debug.txt', 'w') as f:
            f.write(traceback.format_exc())


setup()  # initial test envrionment on TC
file_path = TxtUtils('c:\\svc\\svcconfig.ini').get_source().strip().split(':', 1)[1]
print(file_path)
file_path = os.path.dirname(file_path)
ElementlibPath = os.path.join(file_path, 'Configuration\\elementLib.ini')


def getElementMapping(filepath=ElementlibPath):
    """
    # element format:
    # define name:"name":automationid:controltype
    # eg.: OKButton:"OK":Button----->By name
    #      CancelButton:btnCancel:Button----->By automationId
    :param filepath: ElementlibPath
    :return: element
    """
    mappingDict = {}
    lines = TxtUtils(filepath).get_lines()
    for line in lines:
        if line[0] == '#':
            continue
        items = line.strip().split(":", 1)
        mappingDict[items[0]] = items[1]
    return mappingDict


def getElement(name, regex=True, **kwargs):
    # name is defined name, format: defined name:"Name"/AutomationId:ControlType
    elementId = getElementMapping()[name].split(':')[0]
    controltype = getElementMapping()[name].split(':')[1].upper()
    if elementId.startswith('"') and elementId.endswith('"'):
        if regex:
            return getElementByType(controltype, RegexName=elementId.replace('"', ''), **kwargs)
        else:
            return getElementByType(controltype, Name=elementId.replace('"', ''), **kwargs)
    else:
        return getElementByType(controltype, AutomationId=elementId, **kwargs)


class CommonUtils(QAUtils):
    @staticmethod
    def SwitchToUser():
        QAUtils.SwitchUser("User", "User", "")

    @staticmethod
    def SwitchToAdmin():
        logon_user = platform.version()
        if logon_user.split(".")[0] == "10":
            QAUtils.SwitchUser("Admin", "Admin", "")
        else:
            QAUtils.SwitchUser("Administrator", "Administrator", "")

    @staticmethod
    def launchFromControl():
        QAUtils.LaunchAppFromControl("HP Easy Shell Configuration")

    @staticmethod
    def launchFromPath():
        QAUtils.LaunchAppFromFile("C:\\Program Files\\HP\\HP Easy Shell\\HPEasyShellConfig.exe")

    @staticmethod
    def install(path):
        os.system('msiexec.exe /q /i {}'.format(path))
