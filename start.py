import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from Test_Scripts.easyshell import *
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, colors

error_font = Font(color=colors.RED)


def clear_runtime_folder(path):
    arry = []
    dirs = os.listdir(path)
    for i in dirs:
        if '_MEI' in i:
            arry.append(i)
            os.system('rd /s /q {}\\{}'.format(path, i))


class Test:
    def __init__(self):
        easyshelltest = EasyShellTest()
        self.path = easyshelltest.path
        self.result = True  # sigle test case result
        self.casepath = easyshelltest.casepath
        self.logpath = easyshelltest.log_path
        self.misc = easyshelltest.misc
        self.testset = easyshelltest.testset
        self.testing = easyshelltest.testing

    def getAttachment(self, attachmentFilePath):
        attachment = MIMEText(open(attachmentFilePath, 'rb').read(), 'base64', 'utf-8')
        attachment["Content-Type"] = 'application/octet-stream'
        attachment["Content-Disposition"] = 'attachment;filename=%s' % os.path.basename(attachmentFilePath)
        return attachment

    def sendMail(self, recipient, subject, text, *attachmentFilePaths):
        mailUser = "AutoTest<AutoTest@hp.com>"
        msg = MIMEMultipart('related')
        msg['From'] = mailUser
        msg['To'] = ','.join(recipient)
        msg['Subject'] = subject  # "AddonCatalog check result"
        msg.attach(MIMEText(text, 'html', 'utf-8'))

        for attachmentFilePath in attachmentFilePaths:
            msg.attach(self.getAttachment(attachmentFilePath))
        try:
            mailServer = smtplib.SMTP(host='15.73.212.81', port=25)
            mailServer.ehlo()
            mailServer.starttls()
            mailServer.ehlo()
            mailServer.sendmail(mailUser, recipient, msg.as_string())
            mailServer.close()
            print("Sent email to %s success" % recipient)
        except:
            print("sent email fail~~")
            EasyShellTest().Logfile('send mail fail:\n {}'.format(traceback.format_exc()))

    def test(self):
        if not os.path.exists(os.path.join(self.path, 'flag.txt')):
            os.system("echo testing>{}".format(os.path.join(self.path, 'flag.txt')))
        with open(os.path.join(self.path, 'flag.txt'), 'r') as f:
            src = f.read().upper()
            if "TEST FINISHED" in src:
                print("begin to return")
                return "test finished"
        # ---------Clear template folder by services in volumn C:
        dirs = os.listdir('c:\\')
        for folder in dirs:
            if "_MEI" in folder:
                os.system('rd /s /q c:\\{}'.format(folder))
        # ------------------------------------
        wb = load_workbook(self.testset)
        sheets = wb.sheetnames  # 获得表单名字
        ws = wb[sheets[0]]
        rows = ws.max_row
        for i in range(2, rows + 1):
            self.result = True
            testName = ws.cell(row=i, column=1).value
            result = ws.cell(row=i, column=2).value
            if testName is None:
                break
            if '#' in testName:
                continue
            if result == 'PASS' or result == 'FAIL' or result == 'N/A':
                continue
            else:
                if not os.path.exists(os.path.join(self.testing, '{}.xlsx'.format(testName))):
                    os.system('copy {} {}'.format(os.path.join(self.casepath, '{}.xlsx'.format(testName)),
                                                  os.path.join(self.testing, '{}.xlsx'.format(testName))))
                self.runTestcase(os.path.join(self.testing, '{}.xlsx'.format(testName)))
            if self.result:
                ws.cell(row=i, column=2).value = "PASS"
                wb.save(self.testset)
            else:
                ws.cell(row=i, column=2).value = "FAIL"
                ws.cell(row=i, column=2).font = error_font
                wb.save(self.testset)
        txt = "Attachment is HP EasyShell Test Result"
        subject = "HP EasyShell Test Result"
        mail_list = ["balance.cheng@hp.com"]
        self.sendMail(mail_list, subject, txt, self.testset)
        os.system("echo Test Finished>{}".format(os.path.join(self.path, 'flag.txt')))


    def runTestcase(self, name):
        print('Begin run test case: {}'.format(name))
        wb = load_workbook(name)
        sheets = wb.sheetnames  # 获得表单名字
        ws = wb[sheets[0]]
        rows = ws.max_row
        for i in range(2, rows + 1):
            checkPoint = ws.cell(row=i, column=1).value
            result = ws.cell(row=i, column=4).value
            command = ws.cell(row=i, column=2).value
            value = ws.cell(row=i, column=3).value
            if checkPoint is None:
                break
            if '#' in checkPoint:
                continue
            if result == 'PASS':
                continue
            if result == 'FAIL':
                self.result = False
                continue
            if checkPoint.upper() == "Y":
                rs = eval(command)
                print(rs)
                if value is None:
                    if rs:
                        ws.cell(row=i, column=4).value = "PASS"
                        wb.save(name)
                        EasyShellTest().Logfile("--->[Pass]:{} check".format(command))
                    else:
                        self.result = False
                        ws.cell(row=i, column=4).value = "FAIL"
                        ws.cell(row=i, column=4).font = error_font
                        wb.save(name)
                        EasyShellTest().Logfile("--->[Fail]:{} check, Expect:{},Actual:{}".format(command, "True", rs))
                else:
                    if '$NOT' in str(value).upper():
                        value = str(value).upper().replace('$NOT', '').strip()
                        if str(rs).upper() != value:
                            ws.cell(row=i, column=4).value = "PASS"
                            wb.save(name)
                            EasyShellTest().Logfile("--->[Pass]:{} check".format(command))
                        else:
                            self.result = False
                            ws.cell(row=i, column=4).value = "FAIL"
                            ws.cell(row=i, column=4).font = error_font
                            wb.save(name)
                            EasyShellTest().Logfile(
                                "--->[Fail]:{} check, Expect:{},Actual:{}".format(command, value, rs))
                    elif '$BETWEEN' in str(value).upper():
                        value = str(value).upper().replace('$BETWEEN', '').strip()
                        rs = int(rs)
                        v1, v2 = value.split(',')
                        v1 = int(v1.strip())
                        v2 = int(v2.strip())
                        if v1 < rs < v2:
                            ws.cell(row=i, column=4).value = "PASS"
                            wb.save(name)
                            EasyShellTest().Logfile("--->[Pass]:{} check".format(command))
                        else:
                            self.result = False
                            ws.cell(row=i, column=4).value = "FAIL"
                            ws.cell(row=i, column=4).font = error_font
                            wb.save(name)
                            EasyShellTest().Logfile(
                                "--->[Fail]:{} check, Expect:{},Actual:{}".format(command, value, rs))
                    else:
                        if str(rs).upper() == str(value).upper():
                            ws.cell(row=i, column=4).value = "PASS"
                            wb.save(name)
                            EasyShellTest().Logfile("--->[Pass]:{} check".format(command))
                        else:
                            self.result = False
                            ws.cell(row=i, column=4).value = "FAIL"
                            ws.cell(row=i, column=4).font = error_font
                            wb.save(name)
                            EasyShellTest().Logfile(
                                "--->[Fail]:{} check, Expect:{},Actual:{}".format(command, value, rs))
            elif checkPoint.upper() == 'N':
                if 'Reboot' in str(command):
                    ws.cell(row=i, column=4).value = "PASS"
                    wb.save(name)
                    clear_runtime_folder("C:")
                    exec(command)
                else:
                    print(command)
                    exec(command)
                    ws.cell(row=i, column=4).value = "PASS"
                    wb.save(name)
                    print('save no check')
            else:
                continue
        wb.save(name)


if __name__ == '__main__':
    Test().test()
