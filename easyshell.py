from openpyxl import load_workbook
from EasyShell_Lib import *
import ruamel.yaml as yaml
import time, os, sys


def ClearContent(lenth=50):
    for i in range(lenth):
        QATools.SendKey(QATools.Keys().VK_DELETE, 0.01)


class EasyShellTest:
    """
    profile is a test configuration name, it is a dict for app options, load from easyshellcreate.yaml
    """

    def __init__(self):
        if not os.path.exists('c:\\svc'):
            os.mkdir("C:\\svc")
        if not os.path.exists("C:\\svc\\easyshell"):
            os.mkdir("C:\\svc\\easyshell")
        # self.path = ""
        self.path = "c:\\svc\\easyshell"
        self.result = True
        self.casepath = os.path.join(self.path, "testcases\\EditApp.xlsx")

    # ----------------- Administrator Settings creation and modification ----------
    def ModifySettings(self, profile):
        flag = True
        with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
            print('Open yaml..')
            content = yaml.safe_load(f)
            appPath = content['appPath']['easyShellPath']
            test = content['userInterface'][profile]
            Common.LaunchAppFromFile(appPath)
            time.sleep(5)
            for i in test:
                name = i.split(":")[0].strip()
                status = i.split(":")[1].strip()
                if status == 'ON':
                    try:
                        UserSettings_Dict[name].Enable()
                    except:
                        flag = False
                        self.Logfile('[Fail]Button {} Enable'.format(name))
                elif status == 'OFF':
                    try:
                        UserSettings_Dict[name].Disable()
                    except:
                        flag = False
                        self.Logfile('[Fail]Button {} Disable'.format(name))
                else:
                    flag = False
                    self.Logfile('[Fail]Button {} Status in profile is not Correct!'.format(name))
        ButtonList.APPLY().Click()
        return flag

    def CheckSettings(self, profile):
        flag = True
        with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
            print('Open yaml..')
            content = yaml.safe_load(f)
            appPath = content['appPath']['easyShellPath']
            test = content['userInterface'][profile]
            swTitles = True
            swBrowser = True
            swPower = True
            swTasksw = True
            swSettings = True
        for i in test:
            """
            1. 没有测试总开关关闭但是子开关开启时的状态
            2. Wifi 只测试链接窗口是否弹出，没有测试wifi连接功能
            """
            name = i.split(":")[0].strip()
            status = i.split(":")[1].strip()
            # /////// All the icons and texts that can be display or not in the main wnd ////////////
            if status == "ON":
                # ///////判断主按钮是否关闭，如果OFF,子按钮不再检查
                if not swTitles:
                    if name in ['DisplayApp', 'DisplayConnections', 'DisplayStoreFront', 'DisplayWebsites']:
                        continue
                if not swBrowser:
                    if name in ['DisplayAddress', 'DisplayNavigation', 'DisplayHome']:
                        continue
                if not swPower:
                    if name in ['AllowLock', 'AllowLogoff', 'AllowRestart', 'AllowShutDown']:
                        continue
                if not swTasksw:
                    if name in ['Permanent', 'DisplaySwitcherTime', 'DisplayBattery', 'DisplayCellular',
                                'DisplaySound', 'DisplasySoundIconInteraction', 'DisplayWifi', 'DisplayWifiInterAction',
                                'DisplayWriteFilter', 'DisplayWriteFilterInteraction']:
                        continue
                if not swSettings:
                    if name in ['AllowMouse', 'AllowKeyboard', 'AllowDisplay', 'AllowSound', 'AllowRegion',
                                'AllowNetworkConn', 'AllowDateTime', 'AllowEasyAccess', 'AllowIEProperty',
                                'AllowWifiConfig']:
                        continue
                # ///////////////////////////////////
                if name == 'DisplayTitle':
                    if UserKiosk_Dict['UserTitles'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'DisplayBrowser':
                    if UserKiosk_Dict['UserBrowser'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'DisplayAdmin':
                    if UserKiosk_Dict['UserAdmin'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'DisplayPower':
                    if UserKiosk_Dict['UserPower'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'EnableTaskSwitcher':
                    if UserKiosk_Dict['TaskSwitcher'].Exists(0, 0):
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'AllowUserSetting':
                    if UserKiosk_Dict['UserSettings'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                # ////////// item for Titles //////////////////////////////////////
                if name == 'DisplayApp':
                    UserKiosk_Dict['UserTitles'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['UserApp'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'DisplayConnections':
                    UserKiosk_Dict['UserTitles'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['UserConnection'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'DisplayStoreFront':
                    UserKiosk_Dict['UserTitles'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['UserStoreFront'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'DisplayWebsites':
                    UserKiosk_Dict['UserTitles'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['UserWebsites'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                # ////////item for Web browser ///////////////////////////////
                if name == 'DisplayAddress':
                    UserKiosk_Dict['UserBrowser'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['AddressBar'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'DisplayHome':
                    UserKiosk_Dict['UserBrowser'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['WebHome'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                # ------------------------Item for Admin power----------------------------------------
                if name == 'AllowLock':
                    UserKiosk_Dict['UserPower'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['Lock'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'AllowLogoff':
                    UserKiosk_Dict['UserPower'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['Logoff'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'AllowRestart':
                    UserKiosk_Dict['UserPower'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['Restart'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'AllowShutDown':
                    UserKiosk_Dict['UserPower'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['Shutdown'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                # ----------------- Virtual keyboard --------
                if name == 'DisplayVKeyboard':
                    if UserKiosk_Dict['UserKeyBoard'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                # ---Label that display mac/time/version... at the bottom of UI -----
                if name == 'DisplayTime':
                    if UserKiosk_Dict['Time'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                        real_time = QATools.getLocalTime('%H:%M')
                        show_time = UserKiosk_Dict['Time'].Name
                        if real_time.slip(':')[0] in show_time:
                            self.Logfile("-->[PASS]: {} real time format".format(name))
                        else:
                            self.Logfile("-->[Fail]: {} real time format".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'DisplayIP':
                    if UserKiosk_Dict['IPAddr'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                        real_ip = QATools.getNetInfo()['IP']
                        show_ip = UserKiosk_Dict['IPAddr']
                        if real_ip == show_ip:
                            self.Logfile("-->[PASS]: {} real IP".format(name))
                        else:
                            self.Logfile("-->[Fail]: {} real IP".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'DisplayMAC':
                    if UserKiosk_Dict['MACAddr'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                        real_mac = QATools.getNetInfo()['MAC']
                        show_mac = UserKiosk_Dict['MACAddr']
                        if real_mac == show_mac:
                            self.Logfile("-->[PASS]: {} real mac".format(name))
                        else:
                            self.Logfile("-->[Fail]: {} real mac".format(name))

                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                # --------------- Task Switcher ----------------------------------
                if name == 'DisplaySwitcherTime':
                    if UserKiosk_Dict['SwitcherTime'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'DisplaySound':
                    if UserKiosk_Dict['SoundIcon'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'DisplaySoundIconInteraction':
                    """
                    判断sound interaction 是否工作
                    1. 判断当前音量大小，大于80时降低音量操作，小于80时增加音量操作
                    2. 通过鼠标键盘两种操作来测试功能
                    """
                    if UserKiosk_Dict['SoundIcon'].IsShown():
                        UserKiosk_Dict['SoundIcon'].Click()
                        currentVol = ButtonList.SoundAdjust().AccessibleCurrentValue()
                        if int(currentVol) < 80:
                            ButtonList.SoundAdjustBar().Drag(10, 0)
                            tempVol = ButtonList.SoundAdjust().AccessibleCurrentValue()
                            if currentVol != tempVol:
                                self.Logfile("[PASS]: {} Sound Adjust by Mouse".format(name))
                            else:
                                flag = False
                                self.Logfile("[FAIL]: {} Sound Adjust by Mouse".format(name))
                            QATools.SendKey(QATools.Keys().VK_RIGHT)
                            finalVol = ButtonList.SoundAdjust().AccessibleCurrentValue()
                            if finalVol != tempVol:
                                self.Logfile("[PASS]: {} Sound Adjust by Keyboard".format(name))
                            else:
                                flag = False
                                self.Logfile("[FAIL]: {} Sound Adjust by Keyboard".format(name))
                        else:
                            ButtonList.SoundAdjustBar().Drag(-10, 0)
                            tempVol = ButtonList.SoundAdjust().AccessibleCurrentValue()
                            if currentVol != tempVol:
                                self.Logfile("[PASS]: {} Sound Adjust by Mouse".format(name))
                            else:
                                flag = False
                                self.Logfile("[FAIL]: {} Sound Adjust by Mouse".format(name))
                            QATools.SendKey(QATools.Keys().VK_LEFT)
                            finalVol = ButtonList.SoundAdjust().AccessibleCurrentValue()
                            if finalVol != tempVol:
                                self.Logfile("[PASS]: {} Sound Adjust by Keyboard".format(name))
                            else:
                                flag = False
                                self.Logfile("[FAIL]: {} Sound Adjust by Keyboard".format(name))
                    else:
                        continue
                if name == 'DisplayWifi':
                    if UserKiosk_Dict['WifiIcon'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'DisplayWifiInterAction':
                    if UserKiosk_Dict['WifiIcon'].IsShown():
                        UserKiosk_Dict['WifiIcon'].Click()
                        time.sleep(3)
                        if Common.WifiSelection.Exists(0, 0):
                            self.Logfile("[PASS]: {} is shown".format(name))
                        else:
                            flag = False
                            self.Logfile("[Fail]: {} is not shown".format(name))
                    else:
                        continue
                if name == 'DisplayWriteFilter':
                    if UserKiosk_Dict['HPWMIcon'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                # ------------No Enable Network status notification -----------
                # ------------No Hide HP Easy Shell during Session ------------
                # -----------No Custom Background -----------------------------
                # ------------User Settings -----------------------------------
                if name == 'AllowMouse':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['SysMouseIcon'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'AllowKeyboard':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['SysKeyboardIcon'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'AllowDisplay':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['SysDisplayIcon'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'AllowSound':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['SysSoundIcon'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'AllowRegion':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['SysRegionIcon'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'AllowNetworkConn':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['SysNetworkConnIcon'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'AllowDateTime':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['SysDateTimeIcon'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'AllowEasyAccess':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['SysEaseAccessCenterIcon'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'AllowIEProperty':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['SysIEIcon'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
                if name == 'AllowWifiConfig':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if UserKiosk_Dict['SysWirelessIcon'].IsShown():
                        self.Logfile("[PASS]: {} is shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is not shown".format(name))
            elif status == "OFF":
                # ///////判断主按钮是否关闭，如果OFF,子按钮不再检查
                if not swTitles:
                    if name in ['DisplayApp', 'DisplayConnections', 'DisplayStoreFront', 'DisplayWebsites']:
                        continue
                if not swBrowser:
                    if name in ['DisplayAddress', 'DisplayNavigation', 'DisplayHome']:
                        continue
                if not swPower:
                    if name in ['AllowLock', 'AllowLogoff', 'AllowRestart', 'AllowShutDown']:
                        continue
                if not swTasksw:
                    if name in ['Permanent', 'DisplaySwitcherTime', 'DisplayBattery', 'DisplayCellular',
                                'DisplaySound', 'DisplasySoundIconInteraction', 'DisplayWifi', 'DisplayWifiInterAction',
                                'DisplayWriteFilter', 'DisplayWriteFilterInteraction']:
                        continue
                if not swSettings:
                    if name in ['AllowMouse', 'AllowKeyboard', 'AllowDisplay', 'AllowSound', 'AllowRegion',
                                'AllowNetworkConn', 'AllowDateTime', 'AllowEasyAccess', 'AllowIEProperty',
                                'AllowWifiConfig']:
                        continue
                # ///////////////////////////////////
                if name == 'DisplayTitle':
                    if not UserKiosk_Dict['UserTitles'].IsShown():
                        swTitles = False
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'DisplayBrowser':
                    if not UserKiosk_Dict['UserBrowser'].IsShown():
                        swBrowser = False
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'DisplayAdmin':
                    if not UserKiosk_Dict['UserAdmin'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'DisplayPower':
                    if not UserKiosk_Dict['UserPower'].IsShown():
                        swPower = False
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'EnableTaskSwitcher':
                    if not UserKiosk_Dict['TaskSwitcher'].Exists(0, 0):
                        swTasksw = False
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'AllowUserSetting':
                    if not UserKiosk_Dict['UserSettings'].IsShown():
                        swSettings = False
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                # ////////// item for Titles //////////////////////////////////////
                if name == 'DisplayApp':
                    UserKiosk_Dict['UserTitles'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['UserApp'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'DisplayConnections':
                    UserKiosk_Dict['UserTitles'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['UserConnection'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'DisplayStoreFront':
                    UserKiosk_Dict['UserTitles'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['UserStoreFront'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'DisplayWebsites':
                    UserKiosk_Dict['UserTitles'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['UserBrowser'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                # ////////item for Web browser ///////////////////////////////
                if name == 'DisplayAddress':
                    UserKiosk_Dict['UserBrowser'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['AddressBar'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'DisplayHome':
                    UserKiosk_Dict['UserBrowser'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['WebHome'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                # ------------------------Item for Admin power----------------------------------------
                if name == 'AllowLock':
                    UserKiosk_Dict['UserPower'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['Lock'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'AllowLogoff':
                    UserKiosk_Dict['UserPower'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['Logoff'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'AllowRestart':
                    UserKiosk_Dict['UserPower'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['Restart'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'AllowShutDown':
                    UserKiosk_Dict['UserPower'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['Shutdown'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                # ----------------- Virtual keyboard --------
                if name == 'DisplayVKeyboard':
                    if not UserKiosk_Dict['UserKeyBoard'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                # ---Label that display mac/time/version... at the bottom of UI -----
                if name == 'DisplayTime':
                    if not UserKiosk_Dict['Time'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'DisplayIP':
                    if not UserKiosk_Dict['IPAddr'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'DisplayMAC':
                    if not UserKiosk_Dict['MACAddr'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                # --------------- Task Switcher ----------------------------------
                if name == 'DisplaySwitcherTime':
                    if not UserKiosk_Dict['SwitcherTime'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'DisplaySound':
                    if not UserKiosk_Dict['SoundIcon'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'DisplaySoundIconInteraction':
                    if not ButtonList.SoundIcon().IsShown():
                        continue
                    else:
                        ButtonList.SoundIcon().Click()
                        if not ButtonList.SoundAdjust().IsShown():
                            self.Logfile("[PASS]: {} is not shown".format(name))
                        else:
                            flag = False
                            self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'DisplayWifi':
                    if not UserKiosk_Dict['WifiIcon'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'DisplayWifiInterAction':
                    if UserKiosk_Dict['WifiIcon'].IsShown():
                        UserKiosk_Dict['WifiIcon'].Click()
                        time.sleep(3)
                        if not Common.WifiSelection.Exists(0, 0):
                            self.Logfile("[PASS]: {} is not shown".format(name))
                        else:
                            flag = False
                            self.Logfile("[Fail]: {} is shown".format(name))
                    else:
                        continue
                if name == 'DisplayWriteFilter':
                    if not UserKiosk_Dict['HPWMIcon'].IsShown():
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                # ------------No Enable Network status notification -----------
                # ------------No Hide HP Easy Shell during Session ------------
                # -----------No Custom Background -----------------------------
                # ------------User Settings -----------------------------------
                if name == 'AllowMouse':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['SysMouseIcon'].Exists(0, 0):
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'AllowKeyboard':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['SysKeyboardIcon'].Exists(0, 0):
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'AllowDisplay':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['SysDisplayIcon'].Exists(0, 0):
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'AllowSound':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['SysSoundIcon'].Exists(0, 0):
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'AllowRegion':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['SysRegionIcon'].Exists(0, 0):
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'AllowNetworkConn':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['SysNetworkConnIcon'].Exists(0, 0):
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'AllowDateTime':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['SysDateTimeIcon'].Exists(0, 0):
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'AllowEasyAccess':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['SysEaseAccessCenterIcon'].Exists(0, 0):
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'AllowIEProperty':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['SysIEIcon'].Exists(0, 0):
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
                if name == 'AllowWifiConfig':
                    UserKiosk_Dict['UserSettings'].Click()
                    time.sleep(1)
                    if not UserKiosk_Dict['SysWirelessIcon'].Exists(0, 0):
                        self.Logfile("[PASS]: {} is not shown".format(name))
                    else:
                        flag = False
                        self.Logfile("[Fail]: {} is shown".format(name))
        return flag

    # ----------------- StoreFront connection creation ----------------------------
    def CreateStoreFront(self, profile):
        with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
            test = yaml.safe_load(f)
            appPath = test['appPath']['easyShellPath']
            test = test['createStoreFront'][profile]
            Name = test["Name"]
            URL = test['URL']
            SelectStore = test['SelectStore']
            StoreName = test['StoreName']
            Launchdelay = test['Launchdelay']
            LogonMethod = test['LogonMethod']
            Username = test['Username']
            Password = test['Password']
            Domain = test['Domain']
            HideDomain = test['HideDomain']
            CustomLogon = test['CustomLogon']
            Autolaunch = test['Autolaunch']
            ConnectionTimeout = test['ConnectionTimeout']
            DesktopToolbar = test['DesktopToolbar']
        try:
            Common.LaunchAppFromFile(appPath)
            for i in range(10):
                if not QAUtil.WindowMethod(RegexName="HP Easy Shell.*").Exists(maxSearchSeconds=1):
                    continue
                else:
                    print('store get window')
                    print(ButtonList.KioskMode().Exists(0, 0))
                    break
            ButtonList.KioskMode().Enable()
            ButtonList.DisplayTitle().Enable()
            ButtonList.DisplayStoreFront().Enable()
            ButtonList.StoreFront().Click()
            if self.StoreUtils(profile, 'Exist'):
                self.StoreUtils(profile, 'Delete')
            ButtonList.StoreFrontAdd().Click()
            time.sleep(5)
            QATools.SendKeys(Name)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKeys(URL)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKeys(LogonMethod)
            QATools.SendKey(QATools.Keys().VK_TAB)
            if HideDomain == 'OFF':
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKeys(Username)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKeys(Password)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKeys(Domain)
            QATools.SendKey(QATools.Keys().VK_TAB)
            if Autolaunch == 'OFF':
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
            if not SelectStore:
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                time.sleep(1)
                QATools.SendKeys(StoreName)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_SPACE)
                time.sleep(3)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
            ClearContent()
            QATools.SendKeys(str(Launchdelay))
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            if CustomLogon == 'None':
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKeys(CustomLogon)
                QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_RIGHT)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            if DesktopToolbar == 'OFF':
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
            ClearContent()
            QATools.SendKeys(str(ConnectionTimeout))
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_SPACE)
            ButtonList.APPLY().Click()
            self.Logfile('[PASS]: View Connection {} Create'.format(Name))
            return True
        except Exception as e:
            self.Logfile("[FAIL]:Storfront {} Create\nErrors:\n{}\n".format(Name, e))
            return False

    # ----------------- VMWare view connection Creation ----------------------------
    def CreateView(self, profile):
        with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
            test = yaml.safe_load(f)
            appPath = test['appPath']['easyShellPath']
            test = test['createView'][profile]
            Name = test["Name"]
            Hostname = test['Hostname']
            Launchdelay = test['Launchdelay']
            Argument = test['Argument']
            Autolaunch = test['Autolaunch']
            Persistent = test['Persistent']
            Layout = test['Layout']
            ConnUSBStartup = test['ConnUSBStartup']
            ConnUSBInsertion = test['ConnUSBInsertion']
            Username = test['Username']
            Password = test['Password']
            Domain = test['Domain']
            DesktopName = test['DesktopName']
        try:
            Common.LaunchAppFromFile(appPath)
            for i in range(10):
                if not QAUtil.WindowMethod(RegexName="HP Easy Shell.*").Exists(maxSearchSeconds=1):
                    continue
                else:
                    print('view get window')
                    break
            ButtonList.KioskMode().Enable()
            ButtonList.DisplayTitle().Enable()
            ButtonList.DisplayConnections().Enable()
            ButtonList.Connections().Click()
            if self.ViewUtils(profile, 'exist'):
                self.ViewUtils(profile, 'Delete')
            ButtonList.VMwareAdd().Click()
            time.sleep(5)
            QATools.SendKeys(Name)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKeys(Hostname)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            ClearContent()
            QATools.SendKeys(str(Launchdelay))
            QATools.SendKey(QATools.Keys().VK_TAB)
            if Argument == 'None' or Argument is None:
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKeys(Argument)
                QATools.SendKey(QATools.Keys().VK_TAB)
            if Autolaunch == 'OFF':
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
            if Persistent == 'OFF':
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_RIGHT)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKeys(Layout)
            QATools.SendKey(QATools.Keys().VK_TAB)
            if ConnUSBStartup == 'OFF':
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
            if ConnUSBInsertion == 'OFF':
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKeys(Username)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKeys(Password)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKeys(Domain)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKeys(DesktopName)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_SPACE)
            ButtonList.APPLY().Click()
            self.Logfile('[PASS]: View Connection {} Create'.format(Name))
            return True
        except Exception as e:
            self.Logfile("[FAIL]: View Connection {} Create\nErrors:\n{}\n".format(Name, e))
            return False

    #   ---------------- Website Creation --------------------------
    def CreateWebsite(self, profile):
        with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
            test = yaml.safe_load(f)
            appPath = test['appPath']['easyShellPath']
            test = test['createWebsites'][profile]
            Name = test["Name"]
            Address = test['Address']
            DefaultHome = test['DefaultHome']
            UseIE = test['UseIE']
            IEFallScreen = test['IEFullScreen']
            EmbedIE = test['EmbedIE']
            AllCloseEmbedIE = test['AllCloseEmbedIE']
            EmbaedPaneName = test['EmbaedPaneName']
            try:
                Common.LaunchAppFromFile(appPath)
                for i in range(10):
                    if not QAUtil.WindowMethod(RegexName="HP Easy Shell.*").Exists(maxSearchSeconds=1):
                        continue
                    else:
                        print('web Get windows')
                        break
                ButtonList.KioskMode().Enable()
                UserSettings_Dict['DisplayTitle'].Enable()
                UserSettings_Dict['DisplayWebsites'].Enable()
                UserSettings_Dict['DisplayBrowser'].Enable()
                UserSettings_Dict['DisplayAddress'].Enable()
                UserSettings_Dict['DisplayNavigation'].Enable()
                UserSettings_Dict['DisplayHome'].Enable()
                ButtonList.WebSites().Click()
                if self.WebUtils(profile, 'Exist'):
                    self.WebUtils(profile, 'Delete')
                ButtonList.WebsiteAdd().Invoke()
                time.sleep(3)
                print(Name)
                QATools.SendKeys(Name)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKeys(Address)
                QATools.SendKey(QATools.Keys().VK_TAB)
                if UseIE:
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    if IEFallScreen:
                        QATools.SendKey(QATools.Keys().VK_SPACE)
                        QATools.SendKey(QATools.Keys().VK_TAB)
                        if EmbedIE:
                            QATools.SendKey(QATools.Keys().VK_SPACE)
                            QATools.SendKey(QATools.Keys().VK_TAB)
                            if AllCloseEmbedIE:
                                QATools.SendKey(QATools.Keys().VK_SPACE)
                                QATools.SendKey(QATools.Keys().VK_TAB)
                                QATools.SendKey(QATools.Keys().VK_TAB)
                                QATools.SendKey(QATools.Keys().VK_SPACE)
                            else:
                                QATools.SendKey(QATools.Keys().VK_TAB)
                                QATools.SendKey(QATools.Keys().VK_TAB)
                                QATools.SendKey(QATools.Keys().VK_SPACE)
                        else:
                            QATools.SendKey(QATools.Keys().VK_TAB)
                            QATools.SendKey(QATools.Keys().VK_TAB)
                            QATools.SendKey(QATools.Keys().VK_SPACE)
                    else:
                        QATools.SendKey(QATools.Keys().VK_TAB)
                        QATools.SendKey(QATools.Keys().VK_TAB)
                        QATools.SendKey(QATools.Keys().VK_SPACE)
                else:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                if DefaultHome:
                    self.WebUtils(profile, 'default')
                ButtonList.APPLY().Click()
                return True
            except Exception as e:
                self.Logfile("[FAIL]:Website {} Create\nErrors:\n{}\n".format(Name, e))
                return False

    # ---------------- Website Check --------------------------------------
    def CheckWebsite(self, profile):
        flag = True
        with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
            test = yaml.safe_load(f)
            test = test['createWebsites'][profile]
            Name = test["Name"]
            Address = test['Address']
            DefaultHome = test['DefaultHome']
            UseIE = test['UseIE']
            IEFullScreen = test['IEFullScreen']
            EmbedIE = test['EmbedIE']
            AllCloseEmbedIE = test['AllCloseEmbedIE']
            EmbaedPaneName = test['EmbaedPaneName']
            if QAUtil.WindowMethod(RegexName='.*- Internet Explorer').Exists(0, 0):
                QAUtil.WindowMethod(RegexName='.*- Internet Explorer').Close()
            ButtonList.UserTitles().Click()
            self.WebUtils(profile, 'launch')
            time.sleep(5)
        if not UseIE:
            if QAUtil.PaneMethod(RegexName=EmbaedPaneName).Exists(0, 0) and UserKiosk_Dict['AddressBar'].IsShown():
                if DefaultHome:
                    UserKiosk_Dict['WebHome'].Click()
                    time.sleep(5)
                    if QAUtil.PaneMethod(RegexName=EmbaedPaneName).Exists(0, 0):
                        self.Logfile("[PASS]: Websites {} Check".format(profile))
                    else:
                        flag = False
                        self.Logfile("[FAIL]: Websites {} Check".format(profile))
                else:
                    self.Logfile("[PASS]: Websites {} Check".format(profile))
            else:
                flag = False
                self.Logfile("[FAIL]: Websites {} Check".format(profile))
        elif UseIE and not IEFullScreen:
            if QAUtil.WindowMethod(RegexName=EmbaedPaneName).Exists(0, 0) and \
                    QAUtil.WindowMethod(RegexName=EmbaedPaneName).PaneControl(AutomationId='41477').Exists():
                self.Logfile("[PASS]: Websites {} Check".format(profile))
            else:
                flag = False
                self.Logfile("[FAIL]: Websites {} Check".format(profile))
        elif UseIE and IEFullScreen and not EmbedIE:
            if QAUtil.WindowMethod(RegexName=EmbaedPaneName).Exists(0, 0) and \
                    not (QAUtil.WindowMethod(RegexName=EmbaedPaneName).PaneControl(AutomationId='41477').Exists(0, 0)):
                self.Logfile("[PASS]: Websites {} Check".format(profile))
            else:
                flag = False
                self.Logfile("[FAIL]: Websites {} Check(110)".format(profile))
        elif UseIE and IEFullScreen and EmbedIE and not AllCloseEmbedIE:
            if QAUtil.PaneMethod(RegexName=EmbaedPaneName).Exists(0, 0) and not (UserKiosk_Dict['AddressBar'].IsShown()) \
                    and not (UserKiosk_Dict['WebIEClose'].IsShown()):
                if DefaultHome:
                    UserKiosk_Dict['WebHome'].Click()
                    time.sleep(5)
                    if QAUtil.PaneMethod(RegexName=EmbaedPaneName).Exists(0, 0):
                        self.Logfile("[PASS]: Websites {} Check".format(profile))
                    else:
                        flag = False
                        self.Logfile("[FAIL]: Websites {} Home Check(1110)".format(profile))
                else:
                    self.Logfile("[PASS]: Websites {} Check".format(profile))
            else:
                flag = False
                self.Logfile("[FAIL]: Websites {} Check(1110)".format(profile))
        elif UseIE and IEFullScreen and EmbedIE and AllCloseEmbedIE:
            if QAUtil.PaneMethod(RegexName=EmbaedPaneName).Exists(0, 0) and not (UserKiosk_Dict['AddressBar'].IsShown()) \
                    and UserKiosk_Dict['WebIEClose'].IsShown():
                if DefaultHome:
                    UserKiosk_Dict['WebHome'].Click()
                    time.sleep(5)
                    if QAUtil.PaneMethod(RegexName=EmbaedPaneName).Exists(0, 0):
                        self.Logfile("[PASS]: Websites {} Check".format(profile))
                    else:
                        flag = False
                        self.Logfile("[FAIL]: Websites {} Check(1111)".format(profile))
                else:
                    self.Logfile("[PASS]: Websites {} Check".format(profile))
            else:
                flag = False
                self.Logfile("[FAIL]: Websites {} Check(1111)".format(profile))
        else:
            flag = False
            self.Logfile("[FAIL]: Websites {} Parameter Error!".format(profile))
        return flag

    def EditModifyWebsite(self, newProfile, oldProfile):
        """
        Make sure Home Default is the same with OldProfile
        """
        with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
            test = yaml.safe_load(f)
            appPath = test['appPath']['easyShellPath']
            new = test['createWebsites'][newProfile]
            old = test['createWebsites'][oldProfile]
            newName = new["Name"]
            newAddress = new['Address']
            newDefaultHome = new['DefaultHome']
            oldDefaultHome = old['DefaultHome']
            newUseIE = new['UseIE']
            oldUseIE = old['UseIE']
            newIEFullScreen = new['IEFullScreen']
            oldIEFullScreen = old['IEFullScreen']
            newEmbedIE = new['EmbedIE']
            oldEmbedIE = old['EmbedIE']
            newAllCloseEmbedIE = new['AllCloseEmbedIE']
            oldAllCloseEmbedIE = old['AllCloseEmbedIE']
        try:
            Common.LaunchAppFromFile(appPath)
            for i in range(10):
                if not QAUtil.WindowMethod(RegexName="HP Easy Shell.*").Exists(maxSearchSeconds=1):
                    continue
                else:
                    QAUtil.WindowMethod(RegexName="HP Easy Shell.*")
                    break
            ButtonList.WebSites().Click()
            self.WebUtils(oldProfile, 'Edit')
            time.sleep(3)
            ClearContent()
            QATools.SendKeys(newName)
            QATools.SendKey(QATools.Keys().VK_TAB)
            ClearContent()
            QATools.SendKeys(newAddress)
            QATools.SendKey(QATools.Keys().VK_TAB)
            if newUseIE != oldUseIE and newUseIE:
                if newUseIE:
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    ButtonList.APPLY().Click()
                    return True
            else:
                if newUseIE:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    ButtonList.APPLY().Click()
                    return True
            if newIEFullScreen != oldIEFullScreen:
                if newIEFullScreen:
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    ButtonList.APPLY().Click()
                    return True
            else:
                if newIEFullScreen:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    ButtonList.APPLY().Click()
                    return True
            if newEmbedIE != oldEmbedIE:
                if newEmbedIE:
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    ButtonList.APPLY().Click()
                    return True
            else:
                if newEmbedIE:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    ButtonList.APPLY().Click()
                    return True
            if newAllCloseEmbedIE != oldAllCloseEmbedIE:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_SPACE)
                ButtonList.APPLY().Click()
                return True
            else:
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_SPACE)
                ButtonList.APPLY().Click()
                return True
        except Exception as e:
            self.Logfile('[Fail]: Edit website {}'.format(oldProfile))
            return False

    # --------------- Applications Creation --------------------
    def CheckApp(self, profile):
        flag = True
        with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
            test = yaml.safe_load(f)
            test = test['createApp'][profile]
            Name = test["Name"]
            Launchdelay = test['Launchdelay']
            Autolaunch = test['Autolaunch']
            Persistent = test['Persistent']
            Maximized = test['Maximized']
            Adminonly = test['Adminonly']
            Hidemissapp = test['Hidemissapp']
            WindowName = test['WindowName']
            print(Name)
            try:
                if Adminonly:
                    if self.AppUtils(profile, 'exist'):
                        flag = False
                        self.Logfile("[Fail]:App {} AdminOnly".format(Name))
                        return flag
                    else:
                        self.Logfile("[PASS]:App {} AdminOnly".format(Name))
                        return flag
                if Hidemissapp:
                    if self.AppUtils(profile, 'exist'):
                        flag = False
                        self.Logfile("[Failed]:App {} Hide Missing App".format(Name))
                        return flag
                    else:
                        self.Logfile("[PASS]:App {} Hide Missing App".format(Name))
                        return flag
                if Autolaunch == 0:
                    self.AppUtils(profile, "launch")
                    if Launchdelay == "None" or Launchdelay is None:
                        for i in range(5):
                            if QAUtil.WindowMethod(RegexName=WindowName).Exists(0, 0):
                                # self.Logfile("[PASS]:App {} (No AutoLaunch)Manual Launch".format(Name))
                                break
                            else:
                                continue
                        if not QAUtil.WindowMethod(RegexName=WindowName).Exists(0, 0):
                            flag = False
                            self.Logfile("[Failed]:App {} Manual Launch".format(Name))
                            return flag
                    else:
                        time.sleep(3)
                        if QAUtil.WindowMethod(RegexName=WindowName).Exists(0, 0):
                            self.Logfile("[Failed]:APP {} Launch Delay".format(Name))
                            flag = False
                        else:
                            self.Logfile("[PASS]:APP {} Launch Delay".format(Name))
                        time.sleep(20)
                else:
                    if Launchdelay != "None" or Launchdelay is not None:
                        time.sleep(20)
                    if not QAUtil.WindowMethod(RegexName=WindowName).Exists(0, 0):
                        flag = False
                        self.Logfile("[Failed]:App {} AutoLaunch".format(Name))
                        return flag
                    else:
                        self.Logfile("[PASS]:APP {} AutoLaunch".format(Name))
                        self.Logfile("[PASS]:APP {} Launch Delay".format(Name))
                if Maximized:
                    if QAUtil.WindowMethod(RegexName=WindowName).IsMaximize():
                        self.Logfile("[PASS]:App {} Maximized".format(Name))
                    else:
                        self.Logfile("[Failed]:App {} Maximized".format(Name))
                        flag = False
                if Persistent:
                    QAUtil.WindowMethod(RegexName=WindowName).Close()
                    time.sleep(3)
                    if Launchdelay == "None" or Launchdelay is None:
                        if not QAUtil.WindowMethod(RegexName=WindowName).Exists(0, 0):
                            flag = False
                            self.Logfile("[Failed]:App {} Persistent".format(Name))
                            return flag
                        else:
                            self.Logfile("[PASS]:App {} Persistent".format(Name))
                            # self.Logfile("[PASS]:App {} Not AutoDelay".format(Name))
                    else:
                        time.sleep(5)
                        if QAUtil.WindowMethod(RegexName=WindowName).Exists(0, 0):
                            flag = False
                            self.Logfile("[Failed]:APP {} AutoDelay".format(Name))
                else:
                    QAUtil.WindowMethod(RegexName=WindowName).Close()
                    time.sleep(8)
                    if QAUtil.WindowMethod(RegexName=WindowName).Exists(0, 0):
                        flag = False
                        self.Logfile("[Failed]:App {} No Persistent".format(Name))
                        return flag
                    else:
                        self.Logfile("[PASS]:App {} No Persistent".format(Name))
                return flag
            except Exception as ex:
                self.Logfile("[Failed]:App {} App Check\nErrors:\n{}\n".format(Name, ex))
                return False

    def CreateApp(self, profile):
        """
        profile is a test configuration name, it is a dict for app options, load from easyshellcreate.yaml
        """
        with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
            content = yaml.safe_load(f)
            appPath = content['appPath']['easyShellPath']
            test = content['createApp'][profile]
            Name = test["Name"]
            Paths = test['Path']
            for i in Paths:
                if os.path.exists(i):
                    Path = i
                    break
                else:
                    Path = "ErrorPath"
            Argument = test['Argument']
            Launchdelay = test['Launchdelay']
            Autolaunch = test['Autolaunch']
            Persistent = test['Persistent']
            Maximized = test['Maximized']
            Adminonly = test['Adminonly']
            Hidemissapp = test['Hidemissapp']
            try:
                Common.LaunchAppFromFile(appPath)
                for i in range(10):
                    if not QAUtil.WindowMethod(RegexName="HP Easy Shell.*").Exists(maxSearchSeconds=1):
                        continue
                    else:
                        print('app get window')
                        break
                ButtonList.KioskMode().Enable()
                ButtonList.DisplayTitle().Enable()
                ButtonList.DisplayApp().Enable()
                ButtonList.Applications().Click()
                if self.AppUtils(profile, 'Exist'):
                    self.AppUtils(profile, 'Delete')
                ButtonList.ApplicationAdd().Click()
                QATools.Wait(5)
                QATools.SendKeys(Name)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKeys(Path)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_TAB)
                if Argument == "None" or Argument is None:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    QATools.SendKeys(Argument)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                if Launchdelay == "None" or Launchdelay is None:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    if not type(Launchdelay) == int:
                        QATools.SendKeys(Launchdelay)
                        QATools.SendKey(QATools.Keys().VK_TAB)
                    else:
                        QATools.SendKeys(str(Launchdelay))
                        QATools.SendKey(QATools.Keys().VK_TAB)
                # QATools.SendKey(QATools.Keys().VK_TAB) #Exit action logoff and lock and shutdown cannot automate test
                # QATools.SendKeys(Exitaction)
                QATools.SendKey(QATools.Keys().VK_TAB)
                if Autolaunch:
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                if Persistent == 0:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                if Maximized == 0:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                if Adminonly == 0:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                if Hidemissapp == 0:
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    QATools.SendKey(QATools.Keys().VK_SPACE)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_TAB)
                QATools.SendKey(QATools.Keys().VK_SPACE)
                ButtonList.APPLY().Click()
                self.Logfile("[PASS]:App {} Create".format(Name))
                return True
            except Exception as ex:
                self.Logfile("[FAIL]:App {} Create\nErrors:\n{}\n".format(Name, ex))
                return False

    def EditModifyApp(self, newProfile, oldProfile):
        try:
            with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
                test = yaml.safe_load(f)
                appPath = test['appPath']['easyShellPath']
                old = test['createApp'][oldProfile]
                oldPaths = old['Path']
                oldPath = ''
                for path in oldPaths:
                    if os.path.exists(path):
                        oldPath = path
                        break
                if oldPath == '':
                    oldPath = oldPaths[0]
                oldAutolaunch = old['Autolaunch']
                oldPersistent = old['Persistent']
                oldMaximized = old['Maximized']
                oldAdminonly = old['Adminonly']
                oldHidemissapp = old['Hidemissapp']
                new = test['createApp'][newProfile]
                newName = new["Name"]
                newPaths = new['Path']
                newPath = ''
                for path in newPaths:
                    if os.path.exists(path):
                        newPath = path
                        break
                if newPath == '':
                    newPath = newPaths[0]
                newArgument = new['Argument']
                newAutolaunch = new['Autolaunch']
                newPersistent = new['Persistent']
                newMaximized = new['Maximized']
                newAdminonly = new['Adminonly']
                newHidemissapp = new['Hidemissapp']
                newLaunchdelay = new['Launchdelay']
            Common.LaunchAppFromFile(appPath)
            # Wait HP easy shell launch
            time.sleep(3)
            for i in range(10):
                if not QAUtil.WindowMethod(RegexName="HP Easy Shell.*").Exists(maxSearchSeconds=1):
                    continue
                else:
                    QAUtil.WindowMethod(RegexName="HP Easy Shell.*")
                    break
            ButtonList.Applications().Click()
            # Modify setting//////////////////////////////////////
            self.AppUtils(oldProfile, 'edit')
            time.sleep(3)
            ClearContent()
            QATools.SendKeys(newName)
            QATools.SendKey(QATools.Keys().VK_TAB)
            ClearContent(len(oldPath) + 5)
            QATools.SendKeys(newPath)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            ClearContent()
            if newArgument == 'None' or newArgument is None:
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKeys(newArgument)
                QATools.SendKey(QATools.Keys().VK_TAB)
            ClearContent()
            if newLaunchdelay == "None" or newLaunchdelay is None:
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                if not type(newLaunchdelay) == int:
                    QATools.SendKeys(newLaunchdelay)
                    QATools.SendKey(QATools.Keys().VK_TAB)
                else:
                    QATools.SendKeys(str(newLaunchdelay))
                    QATools.SendKey(QATools.Keys().VK_TAB)
            # QATools.SendKey(QATools.Keys().VK_TAB) #Exit action logoff and lock and shutdown cannot automate test
            # QATools.SendKeys(Exitaction)
            QATools.SendKey(QATools.Keys().VK_TAB)
            if oldAutolaunch == newAutolaunch:
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
            if oldPersistent == newPersistent:
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
            if oldMaximized == newMaximized:
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
            if oldAdminonly == newAdminonly:
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
            if oldHidemissapp == newHidemissapp:
                QATools.SendKey(QATools.Keys().VK_TAB)
            else:
                QATools.SendKey(QATools.Keys().VK_SPACE)
                QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_TAB)
            QATools.SendKey(QATools.Keys().VK_SPACE)
            ButtonList.APPLY().Click()
            self.Logfile("[PASS]:App {} Edit\n".format(newName))
            return True
        except Exception as e:
            self.Logfile("[FAIL]:App {} Edit\nErrors:\n{}\n".format(newProfile, e))
            return False

    # ------------------------------ Utils -------------------------------------------
    def Logfile(self, rs):
        with open(os.path.join(self.path, "testlog.txt"), 'a') as f:
            f.write("[{}]:{}\n".format(time.ctime(), rs))

    def WebUtils(self, profile='', op='exist'):
        time.sleep(3)
        with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
            test = yaml.safe_load(f)
            test = test['createWebsites'][profile]
            name = test["Name"]
            if Common.MAIN_WINDOW.TextControl(Name=name).Exists(0, 0):
                print('Name==%s' % name)
                txt = Common.MAIN_WINDOW.TextControl(Name=name)
            else:
                print(name)
                print("didn't get element")
                return False
            websiteControl = txt.GetParentControl().GetParentControl()
            launch = websiteControl.ButtonControl(AutomationId='launchButton2')
            default = websiteControl.ButtonControl(AutoamtionId='homeButton')
            edit = websiteControl.ButtonControl(AutomationId='editButton')
            delete = websiteControl.ButtonControl(AutomationId='deleteButton')
            if op in ["launch", 'LAUNCH', 'Launch']:
                launch.Invoke()
                return True
            elif op in ['edit', 'Edit', 'EDIT']:
                edit.Invoke()
                return True
            elif op in ['delete', 'Delete', 'DELETE']:
                try:
                    delete.Invoke()
                    ButtonList.DeleteYes().Click()
                    ButtonList.APPLY().Click()
                    return True
                except Exception as e:
                    self.Logfile("[FAIL]:App {} Delete\nErrors:\n{}\n".format(name, e))
                    return False
            elif op in ['exist', 'Exist', 'EXIST']:
                return True
            elif op in ['default', 'Default', 'DEFAULT']:
                default.Invoke()
                return True
            else:
                pass

    def StoreUtils(self, profile='', op='exist'):
        time.sleep(3)
        with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
            test = yaml.safe_load(f)
            test = test['createStoreFront'][profile]
            name = test["Name"]
        if Common.MAIN_WINDOW.TextControl(Name=name).Exists(0, 0):
            print('Name==%s' % name)
            txt = Common.MAIN_WINDOW.TextControl(Name=name)
        else:
            print("didn't get element")
            return False
        appControl = txt.GetParentControl().GetParentControl()
        launch = appControl.ButtonControl(AutomationId='LaunchButton')
        edit = appControl.ButtonControl(AutomationId='EditButton')
        delete = appControl.ButtonControl(AutomationId='DeleteButton')
        if op in ["launch", 'LAUNCH', 'Launch']:
            launch.Invoke()
            return True
        elif op in ['edit', 'Edit', 'EDIT']:
            edit.Invoke()
            return True
        elif op in ['delete', 'Delete', 'DELETE']:
            try:
                delete.Invoke()
                ButtonList.DeleteYes().Click()
                ButtonList.APPLY().Click()
                return True
            except Exception as e:
                self.Logfile("[FAIL]:App {} Delete\nErrors:\n{}\n".format(name, e))
                return False
        elif op in ['exist', 'Exist', 'EXIST']:
            return True
        else:
            pass

    def ViewUtils(self, profile='', op='exist'):
        time.sleep(3)
        with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
            test = yaml.safe_load(f)
            test = test['createView'][profile]
            name = test["Name"]
        if Common.MAIN_WINDOW.TextControl(Name=name).Exists(0, 0):
            print('Name==%s' % name)
            txt = Common.MAIN_WINDOW.TextControl(Name=name)
        else:
            print("didn't get element")
            return False
        appControl = txt.GetParentControl().GetParentControl()
        launch = appControl.ButtonControl(AutomationId='launchButton')
        edit = appControl.ButtonControl(AutomationId='editButton')
        delete = appControl.ButtonControl(AutomationId='deleteButton')
        if op in ["launch", 'LAUNCH', 'Launch']:
            launch.Invoke()
            return True
        elif op in ['edit', 'Edit', 'EDIT']:
            edit.Invoke()
            return True
        elif op in ['delete', 'Delete', 'DELETE']:
            try:
                delete.Invoke()
                ButtonList.DeleteYes().Click()
                ButtonList.APPLY().Click()
                return True
            except Exception as e:
                self.Logfile("[FAIL]:App {} Delete\nErrors:\n{}\n".format(name, e))
                return False
        elif op in ['exist', 'Exist', 'EXIST']:
            return True
        else:
            pass

    def AppUtils(self, profile='', op='exist'):
        time.sleep(3)
        with open(os.path.join(self.path, "easyshellCreate.yaml")) as f:
            test = yaml.safe_load(f)
            test = test['createApp'][profile]
            name = test["Name"]
        if Common.MAIN_WINDOW.TextControl(Name=name).Exists(0, 0):
            print('Name==%s' % name)
            txt = Common.MAIN_WINDOW.TextControl(Name=name)
        else:
            print("didn't get element")
            return False
        appControl = txt.GetParentControl().GetParentControl()
        launch = appControl.ButtonControl(AutomationId='launchButton')
        edit = appControl.ButtonControl(AutomationId='editButton')
        delete = appControl.ButtonControl(AutomationId='deleteButton')
        if op in ["launch", 'LAUNCH', 'Launch']:
            launch.Invoke()
            return True
        elif op in ['edit', 'Edit', 'EDIT']:
            edit.Invoke()
            return True
        elif op in ['delete', 'Delete', 'DELETE']:
            try:
                delete.Invoke()
                ButtonList.DeleteYes().Click()
                ButtonList.APPLY().Click()
                return True
            except Exception as e:
                self.Logfile("[FAIL]:App {} Delete\nErrors:\n{}\n".format(name, e))
                return False
        elif op in ['exist', 'Exist', 'EXIST']:
            return True
        else:
            pass

    def test(self):
        wb = load_workbook("C:\\svc\\easyshell\\testset.xlsx")
        sheets = wb.get_sheet_names()  # 获得表单名字
        ws = wb.get_sheet_by_name(sheets[0])
        rows = ws.max_row
        for i in range(2, rows + 1):
            self.result = True
            testName = ws.cell(row=i, column=1).value
            result = ws.cell(row=i, column=2).value
            if result == 'PASS' or result == 'FAIL':
                continue
            else:
                self.casepath = 'C:\\svc\\easyshell\\testcases\\{}.xlsx'.format(testName)
                self.RunTest(self.casepath)
            if self.result:
                ws.cell(row=i, column=2).value = "PASS"
                wb.save("C:\\svc\\easyshell\\testset.xlsx")
            else:
                ws.cell(row=i, column=2).value = "FAIL"
                wb.save("C:\\svc\\easyshell\\testset.xlsx")

    def RunTest(self, name):
        wb = load_workbook(name)
        sheets = wb.get_sheet_names()  # 获得表单名字
        ws = wb.get_sheet_by_name(sheets[0])
        rows = ws.max_row
        print(name, rows)
        for i in range(2, rows + 1):
            checkPoint = ws.cell(row=i, column=1).value
            result = ws.cell(row=i, column=4).value
            command = ws.cell(row=i, column=2).value
            value = ws.cell(row=i, column=3).value
            if result == 'PASS':
                continue
            if result == 'FAIL':
                self.result = False
                continue
            if checkPoint == "Y":
                rs = eval(command)
                print(rs)
                if value is None:
                    if rs:
                        ws.cell(row=i, column=4).value = "PASS"
                        wb.save(self.casepath)
                        self.Logfile("--->[Pass]:{} check".format(command))
                    else:
                        self.result = False
                        ws.cell(row=i, column=4).value = "FAIL"
                        wb.save(self.casepath)
                        self.Logfile("--->[Fail]:{} check, Expect:{},Actual:{}".format(command, "True", rs))
                else:
                    if str(rs).upper() == str(value).upper():
                        ws.cell(row=i, column=4).value = "PASS"
                        wb.save(self.casepath)
                        self.Logfile("--->[Pass]:{} check".format(command))
                    else:
                        self.result = False
                        ws.cell(row=i, column=4).value = "FAIL"
                        wb.save(self.casepath)
                        self.Logfile("--->[Fail]:{} check, Expect:{},Actual:{}".format(command, value, rs))
            else:
                if 'Reboot' in str(command):
                    ws.cell(row=i, column=4).value = "PASS"
                    wb.save(self.casepath)
                    exec(command)
                else:
                    print(command)
                    exec(command)
                    ws.cell(row=i, column=4).value = "PASS"
                    wb.save(self.casepath)
                    print('save no check')
        wb.save(self.casepath)


# Common.addLocalRegValue('SOFTWARE\\HP\\HP Easy Shell', 'KioskModeAdmin', 'True')
# Common.LaunchAppFromFile('C:\\Program Files\\HP\\HP Easy Shell\\HPEasyShellConfig.exe')
# print(QAUtil.TextMethod(RegexName='HP Logon.*').GetParentControl().AccessibleCurrentState())
# print(QAUtil.TextMethod(RegexName='HP Hotkey.*').GetParentControl().AccessibleCurrentState())
# print(ButtonList.AllowSound().AccessibleCurrentState())
# print(ButtonList.AllowRegion().AccessibleCurrentState())
# print(QAUtil.ButtonMethod(Name='Create').AccessibleCurrentState())
# print(QAUtil.ButtonMethod(Name = 'Abort').AccessibleCurrentState())