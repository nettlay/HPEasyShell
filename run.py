import smtplib
import zipfile
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from Test_Scripts.easyshell import *

from openpyxl import load_workbook
from openpyxl.styles import Font, colors

error_font = Font(color=colors.RED)
pass_font = Font(color=colors.GREEN)


def general_report(format='yaml'):
    pass


def zipDir(dirpath, outFullName):
    """
    压缩指定文件夹
    :param dirpath: 目标文件夹路径
    :param outFullName: 压缩文件保存路径+xxxx.zip
    :return: 无
    """
    zip = zipfile.ZipFile(outFullName, "w", zipfile.ZIP_DEFLATED)
    for path, dirnames, filenames in os.walk(dirpath):
        # 去掉目标跟路径，只对目标文件夹下边的文件及文件夹进行压缩
        fpath = path.replace(dirpath, '')
        for filename in filenames:
            zip.write(os.path.join(path, filename), os.path.join(fpath, filename))
    zip.close()


def clear_runtime_folder(path):
    arry = []
    dirs = os.listdir(path)
    for i in dirs:
        if '_MEI' in i:
            arry.append(i)
            os.system('rd /s /q {}\\{}'.format(path, i))


def getPN():
    reg = CommonLib.Reg_Utils()
    key = reg.open("hardware\\description\\system\\bios")
    value = reg.get_value(key, "SystemProductName")
    reg.close(key)
    return value[0]


class Test:
    def __init__(self):
        self.result = {}
        easyshelltest = EasyShellTest()
        self.path = easyshelltest.path
        self.single_case_result = True  # sigle test case result
        self.casepath = easyshelltest.casepath
        self.logpath = easyshelltest.log_path
        if not os.path.exists(self.logpath):
            os.mkdir(self.logpath)
        self.config = easyshelltest.config_path
        self.misc = easyshelltest.misc
        self.testset = easyshelltest.testset
        self.testing = easyshelltest.testing
        self.rs_yml = CommonLib.YmlUtils(os.path.join(self.logpath, 'result.yml'))
        self.case_result = self.rs_yml.get_item()
        self.current_ip = EasyshellLib.CommonUtils.getNetInfo()['IP']

    def getAttachment(self, attachmentFilePath):
        attachment = MIMEText(open(attachmentFilePath, 'rb').read(), 'base64', 'utf-8')
        attachment["Content-Type"] = 'application/octet-stream'
        attachment["Content-Disposition"] = 'attachment;filename=%s' % os.path.basename(attachmentFilePath)
        return attachment

    def sendMail(self, recipient, subject, text, *attachmentFilePaths):
        mailUser = "AutoTest<AutoTest@hp.com>"
        msg = MIMEMultipart('related')
        msg['From'] = mailUser
        msg['To'] = ','.join(recipient)
        msg['Subject'] = subject  # "AddonCatalog check result"
        msg.attach(MIMEText(text, 'html', 'utf-8'))

        for attachmentFilePath in attachmentFilePaths:
            msg.attach(self.getAttachment(attachmentFilePath))
        try:
            mailServer = smtplib.SMTP(host='15.73.212.81', port=25)
            mailServer.ehlo()
            mailServer.starttls()
            mailServer.ehlo()
            mailServer.sendmail(mailUser, recipient, msg.as_string())
            mailServer.close()
            print("Sent email to %s success" % recipient)
        except:
            print("sent email fail~~")
            EasyShellTest().Logfile('send mail fail:\n {}'.format(traceback.format_exc()))

    def test(self):
        if not os.path.exists(os.path.join(self.path, 'flag.txt')):
            os.system("echo testing>{}".format(os.path.join(self.path, 'flag.txt')))
        with open(os.path.join(self.path, 'flag.txt'), 'r') as f:
            src = f.read().upper()
            if "FINISHED" in src:
                print("test finished")
                return "test finished"
        # ---------Clear template folder by services in volumn C:
        dirs = os.listdir('c:\\')
        for folder in dirs:
            if "_MEI" in folder:
                os.system('rd /s /q c:\\{}'.format(folder))
        # ------------------------------------
        if not os.path.exists(self.testing):
            os.mkdir(self.testing)
        wb = load_workbook(self.testset)
        sheets = wb.sheetnames  # 获得表单名字
        ws = wb[sheets[0]]
        rows = ws.max_row
        for i in range(2, rows + 1):
            self.single_case_result = True
            current_case_rs = None
            testName = ws.cell(row=i, column=1).value
            result = ws.cell(row=i, column=2).value
            if testName is None:
                break
            if '#' in testName:
                continue
            if result == 'PASS' or result == 'FAIL' or result == 'N/A':
                continue
            else:
                if not os.path.exists(os.path.join(self.testing, '{}.xlsx'.format(testName))):
                    os.system('copy {} {}'.format(os.path.join(self.casepath, '{}.xlsx'.format(testName)),
                                                  os.path.join(self.testing, '{}.xlsx'.format(testName))))
                # ------ load current test case, if not exist, append one ----------------------
                if not self.case_result:
                    self.rs_yml.write(
                        [{'case_name': testName, 'result': "FAIL", 'steps': [], 'uut_name': self.current_ip}])
                    self.case_result = self.rs_yml.get_item()
                for index in range(len(self.case_result)):
                    if self.case_result[index]['case_name'] == testName:
                        current_case_rs = self.case_result[index]
                if not current_case_rs:
                    self.case_result.extend(
                        [{'case_name': testName, 'result': "Fail", 'steps': [], 'uut_name': self.current_ip}])
                    self.rs_yml.write(self.case_result)
                    current_case_rs = self.case_result[-1]
                # -------------------------------------------------------------------------------
                print('Begin run test case: {}'.format(testName))
                self.runTestcase(os.path.join(self.testing, '{}.xlsx'.format(testName)), current_case_rs)
            if self.single_case_result:
                ws.cell(row=i, column=2).value = "PASS"
                ws.cell(row=i, column=2).font = pass_font
                wb.save(self.testset)
                current_case_rs['result'] = 'pass'
                self.rs_yml.write(self.case_result)
            else:
                ws.cell(row=i, column=2).value = "FAIL"
                ws.cell(row=i, column=2).font = error_font
                wb.save(self.testset)
                current_case_rs['result'] = 'fail'
                self.rs_yml.write(self.case_result)
        zipDir(self.logpath, os.path.join(self.path, 'report.zip'))
        mail_list = []
        with open(os.path.join(self.config, 'mail_list.txt')) as f:
            data = f.readlines()
            for i in data:
                mail_list.append(i.strip())
        version = platform.version().split(".")[0].strip()
        if version == "10":
            os_version = "Wes10"
        else:
            os_version = "Wes7"
        product_name, copy_right, company = General_Test().check_copyright()
        txt = "Attachment is HP EasyShell Test Result</br></br><h4>Name &nbsp&nbsp&nbsp&nbsp&nbsp : {}</br>copyright:" \
              " {}</br>company : {}</h4>".format(product_name, copy_right, company)
        subject = "{} Test Result for {} {}".format(product_name, getPN(), os_version)  # pn: platform
        # mail_list = ["balance.cheng@hp.com"]
        self.sendMail(mail_list, subject, txt, self.testset, os.path.join(self.path, 'report.zip'))
        os.remove(os.path.join(self.path, 'report.zip'))
        os.system("echo Test Finished>{}".format(os.path.join(self.path, 'flag.txt')))

    def runTestcase(self, name, current_case_rs):
        wb = load_workbook(name)
        sheets = wb.sheetnames  # 获得表单名字
        ws = wb[sheets[0]]
        rows = ws.max_row
        for i in range(2, rows + 1):
            checkPoint = ws.cell(row=i, column=1).value
            result = ws.cell(row=i, column=4).value
            command = ws.cell(row=i, column=2).value
            value = ws.cell(row=i, column=3).value
            if checkPoint is None:
                break
            if '#' in checkPoint:
                continue
            if result == 'PASS':
                continue
            if result == 'FAIL':
                self.single_case_result = False
                continue
            if checkPoint.upper() == "DEFINE":
                if ";" in command:
                    command_list = command.split(";")
                    for i in command_list:
                        exec(i)
                else:
                    exec(command)
                continue
            if checkPoint.upper() == "Y":
                print(command)
                rs = eval(command)
                if rs is True:
                    ws.cell(row=i, column=4).value = "PASS"
                    ws.cell(row=i, column=4).font = pass_font
                    wb.save(name)
                    EasyShellTest().Logfile("[Pass]:{} check".format(command))
                    current_case_rs['steps'].append(
                        {"step_name": command, 'actual': "", 'expect': "", 'result': 'PASS',
                         'note': "Test Pass"})
                    self.rs_yml.write(self.case_result)
                elif rs is False:
                    self.single_case_result = False
                    ws.cell(row=i, column=4).value = "FAIL"
                    ws.cell(row=i, column=4).font = error_font
                    wb.save(name)
                    EasyShellTest().Logfile("[Fail]:{} check, Expect:{},Actual:{}".format(command, "True", rs))
                    current_case_rs['steps'].append(
                        {"step_name": command, 'actual': "", 'expect': "", 'result': "FAIL",
                         'note': "Test Fail"})
                    self.rs_yml.write(self.case_result)
                else:
                    step_rs, step_name, step_info, pic_path = rs[0], rs[1], rs[2], rs[3]
                    pic_path = os.path.join(self.logpath, 'img/{}.jpg'.format(pic_path))
                    if value is None:
                        if step_rs.upper() == 'FAIL':
                            self.single_case_result = False
                        ws.cell(row=i, column=4).value = step_rs.upper()
                        ws.cell(row=i, column=4).font = pass_font if step_rs.upper() == "PASS" else error_font
                        wb.save(name)
                        EasyShellTest().Logfile("[{}]:{} check, Actual:{}".format(step_rs, command, rs))
                        current_case_rs['steps'].append(
                            {"step_name": step_name, 'actual': pic_path if step_rs.upper() == "PASS" else "",
                             'expect': "", 'result': step_rs,
                             'note': step_info})
                        self.rs_yml.write(self.case_result)
                    else:
                        if '$NOT' in str(value).upper():
                            value = str(value).upper().replace('$NOT', '').strip()
                            if str(rs).upper() != value:
                                ws.cell(row=i, column=4).value = "PASS"
                                ws.cell(row=i, column=4).font = pass_font
                                wb.save(name)
                                EasyShellTest().Logfile("[Pass]:{} check".format(command))
                                current_case_rs['steps'].append(
                                    {"step_name": step_name, 'actual': '', 'expect': '', 'result': step_rs,
                                     'note': step_info})
                                self.rs_yml.write(self.case_result)
                            else:
                                self.single_case_result = False
                                ws.cell(row=i, column=4).value = "FAIL"
                                ws.cell(row=i, column=4).font = error_font
                                wb.save(name)
                                EasyShellTest().Logfile(
                                    "[Fail]:{} check, Expect:{},Actual:{}".format(command, value, rs))
                                current_case_rs['steps'].append(
                                    {"step_name": step_name, 'actual': pic_path, 'expect': '', 'result': step_rs,
                                     'note': step_info})
                                self.rs_yml.write(self.case_result)
                        elif '$BETWEEN' in str(value).upper():
                            value = str(value).upper().replace('$BETWEEN', '').strip()
                            rs = int(rs)
                            v1, v2 = value.split(',')
                            v1 = int(v1.strip())
                            v2 = int(v2.strip())
                            if v1 < rs < v2:
                                ws.cell(row=i, column=4).value = "PASS"
                                ws.cell(row=i, column=4).font = pass_font
                                wb.save(name)
                                EasyShellTest().Logfile("[Pass]:{} check".format(command))
                                current_case_rs['steps'].append(
                                    {"step_name": step_name, 'actual': '', 'expect': '', 'result': step_rs,
                                     'note': step_info})
                                self.rs_yml.write(self.case_result)
                            else:
                                self.single_case_result = False
                                ws.cell(row=i, column=4).value = "FAIL"
                                ws.cell(row=i, column=4).font = error_font
                                wb.save(name)
                                EasyShellTest().Logfile(
                                    "[Fail]:{} check, Expect:{},Actual:{}".format(command, value, rs))
                                current_case_rs['steps'].append(
                                    {"step_name": step_name, 'actual': pic_path, 'expect': '', 'result': step_rs,
                                     'note': step_info})
                                self.rs_yml.write(self.case_result)
                        else:
                            if str(rs).upper() == str(value).upper():
                                ws.cell(row=i, column=4).value = "PASS"
                                ws.cell(row=i, column=4).font = pass_font
                                wb.save(name)
                                EasyShellTest().Logfile("[Pass]:{} check".format(command))
                                current_case_rs['steps'].append(
                                    {"step_name": step_name, 'actual': '', 'expect': '', 'result': step_rs,
                                     'note': step_info})
                                self.rs_yml.write(self.case_result)
                            else:
                                self.single_case_result = False
                                ws.cell(row=i, column=4).value = "FAIL"
                                ws.cell(row=i, column=4).font = error_font
                                wb.save(name)
                                EasyShellTest().Logfile(
                                    "[Fail]:{} check, Expect:{},Actual:{}".format(command, value, rs))
                                current_case_rs['steps'].append(
                                    {"step_name": step_name, 'actual': pic_path, 'expect': '', 'result': step_rs,
                                     'note': step_info})
                                self.rs_yml.write(self.case_result)
            elif checkPoint.upper() == 'N':
                print(command)
                if 'Reboot' in str(command):
                    ws.cell(row=i, column=4).value = "PASS"
                    ws.cell(row=i, column=4).font = pass_font
                    wb.save(name)
                    clear_runtime_folder("C:")
                    exec(command)
                else:
                    print(command)
                    exec(command)
                    ws.cell(row=i, column=4).value = "PASS"
                    ws.cell(row=i, column=4).font = pass_font
                    wb.save(name)
                    print('save no check')
            else:
                continue
        wb.save(name)


if __name__ == '__main__':
    Test().test()
